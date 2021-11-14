from utils import WORKBOOK_HEADERS
import openpyxl


class TweetDict:
    def __init__(self, dictionary):
        self.dict_ = dictionary

    def to_xlsx(self, filename=None):
        if self.dict_.get("error") is None:
            wb = openpyxl.Workbook()
            ws = wb.active
            for v, i in enumerate(WORKBOOK_HEADERS):
                ws.cell(row=1, column=v + 1).value = i
            max_row = 1
            all_tweets = self.dict_['tweets']
            for i in all_tweets:
                for p in i['result']['tweets']:
                    ws[f'A{max_row + 1}'] = p['created_on']
                    ws[f'B{max_row + 1}'] = p['is_retweet']
                    ws[f'C{max_row + 1}'] = p['is_reply']
                    ws[f'D{max_row + 1}'] = p['tweet_id']
                    ws[f'E{max_row + 1}'] = p['tweet_body']
                    ws[f'F{max_row + 1}'] = p['language']
                    ws[f'G{max_row + 1}'] = p['likes']
                    ws[f'H{max_row + 1}'] = p['retweet_counts']
                    ws[f'I{max_row + 1}'] = p['source']
                    ws[f'N{max_row + 1}'] = p['symbols']
                    media_ = ""
                    if not isinstance(p['media'], str):
                        for media in p['media']:
                            media_ = f"{media_} {media['expanded_url']};"
                        ws[f'J{max_row + 1}'] = media_
                    else:
                        ws[f'J{max_row + 1}'] = p['media']
                    user_mentions = ""
                    if not isinstance(p['user_mentions'], str):
                        for user in p['user_mentions']:
                            user_mentions = f"{user_mentions} {user['screen_name']};"
                        ws[f'K{max_row + 1}'] = user_mentions
                    else:
                        ws[f'K{max_row + 1}'] = p['user_mentions']
                    hashtags = ""
                    if not isinstance(p['hashtags'], str):
                        for tag in p['hashtags']:
                            hashtags = f"{hashtags} {tag['text']};"
                        ws[f'M{max_row + 1}'] = hashtags
                    else:
                        ws[f'M{max_row + 1}'] = p['hashtags']
                    urls = ""
                    if not isinstance(p['urls'], str):
                        for url in p['urls']:
                            urls = f"{urls} {url['expanded_url']};"
                        ws[f'L{max_row + 1}'] = urls
                    else:
                        ws[f'L{max_row + 1}'] = p['urls']
                    max_row = max_row + 1
            if not filename:
                filename = f"tweets.xlsx"
            wb.save(filename)
        else:
            return self.dict_['error']

    def to_dict(self):
        return self.dict_
